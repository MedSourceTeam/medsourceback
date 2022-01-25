import imp
from io import BytesIO
from urllib import request
from wsgiref.util import FileWrapper
from rest_framework import generics
from django.contrib.auth.models import User
from medsource.models import Doctor, user, Nurse, Development
from rest_framework_simplejwt.backends import TokenBackend
from medsourceback import settings
from medsource.serializers import DownloadSerializer
from django.db.models import Q
from openpyxl import load_workbook
from rest_framework.response import Response
from tempfile import NamedTemporaryFile

# Import mimetypes module
import mimetypes
# import os module
import os
# Import HttpResponse module
from django.http.response import HttpResponse


class WorkView(generics.APIView):

    serializer_class = DownloadSerializer

    def get(self, request, *args, **kwargs):

        token = request.META.get('HTTP_AUTHORIZATION')[7:]
        tokenBackend = TokenBackend(algorithm=settings.SIMPLE_JWT['ALGORITHM'])
        valid_data = tokenBackend.decode(token, verify=False)
        queryset = Development.objects.all()

        try:
            person = Doctor.objects.get(user=valid_data["user_id"])

        except:
            person = Nurse.objects.get(user=valid_data["user_id"])

        queryset = queryset.filter(Q(doctor__identification=39074) | Q(nurse__identification=person.identification))

        book = load_workbook(filename="medsource/resources/Formato.xlsx")
        sheet = book.active

        sheet["D4"] = str(person.hospital)
        sheet["D5"] = str(person.user.first_name) + " " + str(person.user.last_name)
        sheet["D8"] = str(person.user.email)
        sheet["J5"] = str(person.identification)

        pos = 16

        for row in queryset:
            sheet["A" + str(pos)] = str(pos-15)
            sheet["B" + str(pos)] = str(row.date)
            sheet["C" + str(pos)] = str(row.patient.full_name)
            sheet["D" + str(pos)] = str(row.patient.identification)
            sheet["E" + str(pos)] = str(row.procedure.name)
            sheet["F" + str(pos)] = str(row.procedure.uvr)
            pos += 1


        '''answer = BytesIO()
        book.save(answer)
        print(answer.read())
        return HttpResponse(answer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')'''



        with open("medsource/resources/tempfile.xlsx", "wb+") as tmp:
            book.save(tmp.name)
            tmp.seek(0)
            return HttpResponse(FileWrapper(tmp), headers={'Content-Disposition': 'attachment; filename="file.xlsx"'},content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        

        #return Response({"data": book.})

        #book.save("medsource/resources/Pagos_" + str(person.identification) + ".xlsx")


        '''BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        # Define text file name
        filename = 'Formato.xlsx'
        # Define the full file path
        filepath = BASE_DIR + '/resources/' + filename
        # Open the file for reading content
        path = open(filepath, 'r')
        # Set the mime type
        mime_type, _ = mimetypes.guess_type(filepath)
        # Set the return value of the HttpResponse
        response = HttpResponse(path, content_type=mime_type)
        # Set the HTTP header for sending to browser
        response['Content-Disposition'] = "attachment; filename=%s" % filename
        # Return the response value
        return response'''
        
        
        # valid_data es un diccionario que contiene toda la info del token, 
        # se puede imprimir para verificar el nombre del campo que trae el id
        # de usuario, normalmente el campo se llama user_id








'''from fileinput import filename
import openpyxl

book = load_workbook(filename="Formato.xlsx")
sheet = book.active'''